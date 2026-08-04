"""
LogiCheck - Accounting & Logistics Telegram Bot
Powered by Google Gemini (Free)
"""

import os
import io
import logging
from datetime import datetime, timedelta
import google.generativeai as genai
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    filters,
    ContextTypes,
)
from telegram.constants import ParseMode

logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Gemini setup ──────────────────────────────────────────────────────────────
genai.configure(api_key=os.environ["GEMINI_API_KEY"])
model = genai.GenerativeModel(
    model_name="gemini-1.5-flash",
    system_instruction="""You are "LogiCheck," an expert AI Accounting & Logistics Assistant for a motor carrier company. Your primary responsibility is to prevent costly accounting errors, flag payment holds, and assist with financial reconciliations.

Your communication style is concise, clear, and direct—optimized for Telegram chat interface. Use simple Markdown (bold with *, bullet points).

When user mentions payment holds, deductions, missing docs (PODs/BOL), rate conflicts — acknowledge and categorize:
Labels: [CARRIER HOLD], [DRIVER DEDUCTION], [VENDOR ALERT], [MISSING DOCS], [RATE CONFLICT]

PRE-PAYMENT AUDIT TRIGGER phrases: "processing payments", "running payouts", "starting settlements", "pre-payment check", "run audit", "audit now"
When triggered: generate full audit report of all active flags.

Keep responses short, bullet points, mobile-friendly.
Never process or approve payments yourself — only audit and flag.
"""
)

# ── In-memory state per user ──────────────────────────────────────────────────
user_sessions: dict[int, dict] = {}

FLAG_KEYWORDS = {
    "[CARRIER HOLD]": "CARRIER HOLD",
    "[DRIVER DEDUCTION]": "DRIVER DEDUCTION",
    "[VENDOR ALERT]": "VENDOR ALERT",
    "[MISSING DOCS]": "MISSING DOCS",
    "[RATE CONFLICT]": "RATE CONFLICT",
}

AUDIT_TRIGGER_KEYWORDS = [
    "processing payments", "running payouts", "starting settlements",
    "pre-payment check", "run audit", "audit now", "payment run",
    "send payments", "batch payment", "process payroll", "run payroll",
]

# ── Main persistent keyboard ──────────────────────────────────────────────────
MAIN_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["📋 Records", "➕ Add New"],
        ["⚠️ Audit", "📊 Export"],
    ],
    resize_keyboard=True,
    is_persistent=True,
)

# ── Session helpers ───────────────────────────────────────────────────────────

def get_session(user_id: int) -> dict:
    if user_id not in user_sessions:
        user_sessions[user_id] = {
            "chat": model.start_chat(history=[]),
            "flags": [],
            "flag_counter": 0,
            "records": [],
            "record_counter": 0,
            "awaiting": None,  # tracks what input we're waiting for
            "export_filter": {},  # tracks export filter state
        }
    return user_sessions[user_id]


def next_record_id(session: dict) -> str:
    session["record_counter"] += 1
    return f"#{session['record_counter']:03d}"


def get_week_label(date: datetime) -> str:
    monday = date - timedelta(days=date.weekday())
    return f"Week of {monday.strftime('%m/%d/%Y')}"


def add_flag(session: dict, category: str, note: str) -> str:
    session["flag_counter"] += 1
    fid = f"F{session['flag_counter']:03d}"
    session["flags"].append({
        "id": fid, "category": category, "note": note,
        "timestamp": datetime.now().strftime("%H:%M"), "resolved": False,
    })
    return fid


def resolve_flag(session: dict, flag_id: str) -> bool:
    for f in session["flags"]:
        if f["id"].upper() == flag_id.upper():
            f["resolved"] = True
            return True
    return False


def build_flags_summary(session: dict) -> str:
    open_flags = [f for f in session["flags"] if not f.get("resolved")]
    resolved = [f for f in session["flags"] if f.get("resolved")]
    if not open_flags and not resolved:
        return "📋 *No flags logged this session.*"
    lines = ["📋 *SESSION FLAGS*\n"]
    if open_flags:
        lines.append("🔴 *OPEN FLAGS:*")
        for f in open_flags:
            lines.append(f"  • `[{f['id']}]` *{f['category']}* — {f['note']} _({f['timestamp']})_")
    else:
        lines.append("🟢 No open flags.")
    if resolved:
        lines.append("\n✅ *RESOLVED:*")
        for f in resolved:
            lines.append(f"  • `[{f['id']}]` {f['category']} — {f['note']}")
    return "\n".join(lines)


def parse_new_record(text: str) -> dict | None:
    parts = [p.strip() for p in text.split("-")]
    if len(parts) < 3:
        return None
    try:
        unit = parts[0] if parts[0] else "N/A"
        name = parts[1]
        amount = float(parts[2].replace("$", "").replace(",", ""))
        ptype = parts[3].lower() if len(parts) > 3 else "deduction"
        payment_method = parts[4] if len(parts) > 4 else "—"
        note = parts[5] if len(parts) > 5 else "—"
        if not name or not amount or not ptype:
            return None
        return {"unit": unit, "name": name, "amount": amount,
                "payment_type": ptype, "payment_method": payment_method, "note": note}
    except (ValueError, IndexError):
        return None


def filter_records(records: list, by: str = None, value: str = None, period: str = None) -> list:
    result = records

    # Filter by field
    if by and value and value.lower() != "all":
        if by == "name":
            result = [r for r in result if r["name"].lower() == value.lower()]
        elif by == "unit":
            result = [r for r in result if str(r["unit"]).lower() == value.lower()]
        elif by == "type":
            result = [r for r in result if r["payment_type"].lower() == value.lower()]

    # Filter by period
    if period:
        now = datetime.now()
        if period == "this_week":
            monday = (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0)
            result = [r for r in result if r["datetime"] >= monday]
        elif period == "last_week":
            monday = (now - timedelta(days=now.weekday() + 7)).replace(hour=0, minute=0, second=0)
            sunday = monday + timedelta(days=6, hours=23, minutes=59)
            result = [r for r in result if monday <= r["datetime"] <= sunday]
        elif period == "this_month":
            start = now.replace(day=1, hour=0, minute=0, second=0)
            result = [r for r in result if r["datetime"] >= start]
        elif period == "last_month":
            first_this = now.replace(day=1, hour=0, minute=0, second=0)
            last_end = first_this - timedelta(seconds=1)
            last_start = last_end.replace(day=1, hour=0, minute=0, second=0)
            result = [r for r in result if last_start <= r["datetime"] <= last_end]

    return result


def build_excel(records: list, title_label: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LogiCheck Records"

    col_widths = {"A": 20, "B": 14, "C": 22, "D": 12, "E": 18, "F": 14, "G": 12, "H": 12, "I": 12, "J": 30, "K": 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"LogiCheck — {title_label}"
    title_cell.font = Font(name="Arial", bold=True, size=13, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = {"A": "Week Of", "B": "Date", "C": "Driver Name", "D": "Unit #",
               "E": "Payment Type", "F": "Amount ($)", "G": "", "H": "", "I": "", "J": "Note", "K": "Payment Method"}
    for col_letter, header in headers.items():
        cell = ws[f"{col_letter}2"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[2].height = 22

    type_colors = {"deduction": "FFF2CC", "reimbursement": "E2EFDA", "bonus": "DDEEFF", "wagehold": "FCE4D6"}
    alt_colors  = {"deduction": "FFF9E6", "reimbursement": "F0F9EC", "bonus": "EEF6FF", "wagehold": "FEF0EB"}
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for i, r in enumerate(records):
        row = i + 3
        ptype = r["payment_type"].lower()
        fill_color = type_colors.get(ptype, "FFFFFF") if i % 2 == 0 else alt_colors.get(ptype, "F9F9F9")
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        dt = r["datetime"]
        values = {"A": get_week_label(dt), "B": dt.strftime("%m/%d/%Y"), "C": r["name"], "D": r["unit"],
                  "E": r["payment_type"].title(), "F": r["amount"], "G": "", "H": "", "I": "", "J": r["note"], "K": r["payment_method"]}
        for col_letter, val in values.items():
            cell = ws[f"{col_letter}{row}"]
            cell.value = val
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border
            if col_letter == "F":
                cell.number_format = '$#,##0.00'
                cell.alignment = center
            elif col_letter in ("A", "B", "D", "E", "K"):
                cell.alignment = center
            else:
                cell.alignment = left
        ws.row_dimensions[row].height = 18

    total_row = len(records) + 3
    ws[f"A{total_row}"].value = "TOTAL"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, size=10)
    ws[f"A{total_row}"].alignment = center
    ws[f"F{total_row}"].value = f"=SUM(F3:F{total_row - 1})" if records else 0
    ws[f"F{total_row}"].font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    ws[f"F{total_row}"].number_format = '$#,##0.00'
    ws[f"F{total_row}"].alignment = center
    ws[f"F{total_row}"].fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def call_gemini(session: dict, user_message: str) -> str:
    is_audit = any(kw in user_message.lower() for kw in AUDIT_TRIGGER_KEYWORDS)
    open_flags = [f for f in session["flags"] if not f.get("resolved")]
    flags_context = ""
    if open_flags:
        flag_list = "\n".join(f"  - [{f['id']}] {f['category']}: {f['note']}" for f in open_flags)
        flags_context = f"\n\n[SYSTEM CONTEXT — Active flags:]\n{flag_list}"
    if is_audit:
        augmented = f"{user_message}{flags_context}\n\n[SYSTEM: PRE-PAYMENT AUDIT triggered.]"
    else:
        augmented = user_message + flags_context
    response = session["chat"].send_message(augmented)
    return response.text


# ── Inline keyboards ──────────────────────────────────────────────────────────

def records_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💸 Deductions", callback_data="rec_deduction"),
         InlineKeyboardButton("💚 Reimbursements", callback_data="rec_reimbursement")],
        [InlineKeyboardButton("⭐ Bonuses", callback_data="rec_bonus"),
         InlineKeyboardButton("🔴 Wage Holds", callback_data="rec_wagehold")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="main_menu")],
    ])


def add_new_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("💸 Deduction", callback_data="add_deduction"),
         InlineKeyboardButton("💚 Reimbursement", callback_data="add_reimbursement")],
        [InlineKeyboardButton("⭐ Bonus", callback_data="add_bonus"),
         InlineKeyboardButton("🔴 Wage Hold", callback_data="add_wagehold")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="main_menu")],
    ])


def audit_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔍 Run Audit", callback_data="audit_run"),
         InlineKeyboardButton("📋 View Flags", callback_data="audit_flags")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="main_menu")],
    ])


def export_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👤 By Name", callback_data="export_by_name"),
         InlineKeyboardButton("🚛 By Unit", callback_data="export_by_unit")],
        [InlineKeyboardButton("📋 By Type", callback_data="export_by_type"),
         InlineKeyboardButton("📅 By Period", callback_data="export_by_period")],
        [InlineKeyboardButton("🔙 Main Menu", callback_data="main_menu")],
    ])


def period_menu(prefix: str):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("This Week", callback_data=f"{prefix}_this_week"),
         InlineKeyboardButton("Last Week", callback_data=f"{prefix}_last_week")],
        [InlineKeyboardButton("This Month", callback_data=f"{prefix}_this_month"),
         InlineKeyboardButton("Last Month", callback_data=f"{prefix}_last_month")],
        [InlineKeyboardButton("All", callback_data=f"{prefix}_all")],
        [InlineKeyboardButton("🔙 Export Menu", callback_data="export_menu")],
    ])


def name_filter_menu(session: dict):
    names = list(set(r["name"] for r in session["records"]))
    buttons = [[InlineKeyboardButton("All", callback_data="fname_all")]]
    row = []
    for name in names:
        row.append(InlineKeyboardButton(name, callback_data=f"fname_{name}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton("🔙 Export Menu", callback_data="export_menu")])
    return InlineKeyboardMarkup(buttons)


def unit_filter_menu(session: dict):
    units = list(set(str(r["unit"]) for r in session["records"]))
    buttons = [[InlineKeyboardButton("All", callback_data="funit_all")]]
    row = []
    for unit in units:
        row.append(InlineKeyboardButton(unit, callback_data=f"funit_{unit}"))
        if len(row) == 3:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton("🔙 Export Menu", callback_data="export_menu")])
    return InlineKeyboardMarkup(buttons)


def type_filter_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("All", callback_data="ftype_all")],
        [InlineKeyboardButton("💸 Deduction", callback_data="ftype_deduction"),
         InlineKeyboardButton("💚 Reimbursement", callback_data="ftype_reimbursement")],
        [InlineKeyboardButton("⭐ Bonus", callback_data="ftype_bonus"),
         InlineKeyboardButton("🔴 Wage Hold", callback_data="ftype_wagehold")],
        [InlineKeyboardButton("🔙 Export Menu", callback_data="export_menu")],
    ])


# ── Send Excel helper ─────────────────────────────────────────────────────────

async def send_excel(message, records: list, label: str):
    if not records:
        await message.reply_text("⚠️ No records found for this filter.", reply_markup=MAIN_KEYBOARD)
        return
    excel_bytes = build_excel(records, label)
    filename = f"LogiCheck_{label.replace(' ', '_').replace('/', '-')}_{datetime.now().strftime('%m%d%Y')}.xlsx"
    await message.reply_document(
        document=io.BytesIO(excel_bytes),
        filename=filename,
        caption=(
            f"📊 *LogiCheck Export*\n"
            f"Filter: *{label}*\n"
            f"Records: *{len(records)}*\n"
            f"Total: *${sum(r['amount'] for r in records):,.2f}*"
        ),
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=MAIN_KEYBOARD,
    )


# ── Command & message handlers ────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    get_session(update.effective_user.id)
    await update.message.reply_text(
        "👋 *LogiCheck Online*\n"
        "Accounting & Logistics Safety Net — Active\n\n"
        "Use the menu below to get started:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=MAIN_KEYBOARD,
    )


async def cmd_resolve(update: Update, context: ContextTypes.DEFAULT_TYPE):
    session = get_session(update.effective_user.id)
    if not context.args:
        await update.message.reply_text("⚠️ Usage: `/resolve F001`", parse_mode=ParseMode.MARKDOWN)
        return
    flag_id = context.args[0].upper()
    if resolve_flag(session, flag_id):
        await update.message.reply_text(f"✅ Flag `{flag_id}` resolved.", parse_mode=ParseMode.MARKDOWN)
    else:
        await update.message.reply_text(f"❌ Flag `{flag_id}` not found.", parse_mode=ParseMode.MARKDOWN)


# ── Callback handler ──────────────────────────────────────────────────────────

async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    session = get_session(query.from_user.id)
    data = query.data

    # ── MAIN MENU ──
    if data == "main_menu":
        await query.message.reply_text("Main Menu:", reply_markup=MAIN_KEYBOARD)

    # ── RECORDS ──
    elif data == "show_records":
        await query.message.reply_text("📋 *Records*\nChoose a category:", parse_mode=ParseMode.MARKDOWN, reply_markup=records_menu())

    elif data.startswith("rec_"):
        ptype = data.replace("rec_", "")
        emoji_map = {"deduction": "💸", "reimbursement": "💚", "bonus": "⭐", "wagehold": "🔴"}
        emoji = emoji_map.get(ptype, "📄")
        filtered = [r for r in session["records"] if r["payment_type"].lower() == ptype]
        if not filtered:
            msg = f"{emoji} *{ptype.title()}s*\n\nNo records yet."
        else:
            lines = [f"{emoji} *{ptype.title()}s* — {len(filtered)} record(s)\n"]
            for r in filtered:
                lines.append(f"`{r['id']}` *{r['name']}* | Unit {r['unit']} | ${r['amount']:,.2f} | {r['payment_method']} | _{r['note']}_")
            lines.append(f"\n💰 *Total: ${sum(r['amount'] for r in filtered):,.2f}*")
            msg = "\n".join(lines)
        await query.message.reply_text(msg, parse_mode=ParseMode.MARKDOWN, reply_markup=records_menu())

    # ── ADD NEW ──
    elif data == "show_add":
        await query.message.reply_text("➕ *Add New Record*\nChoose type:", parse_mode=ParseMode.MARKDOWN, reply_markup=add_new_menu())

    elif data.startswith("add_"):
        ptype = data.replace("add_", "")
        session["awaiting"] = f"new_record_{ptype}"
        await query.message.reply_text(
            f"➕ *New {ptype.title()}*\n\n"
            f"Send in this format:\n"
            f"`Unit - Name - Amount - Type - PaymentMethod - Note`\n\n"
            f"*Required:* Name, Amount\n"
            f"*Example:*\n`2154 - Jama Ahmed - 2500 - {ptype} - Check - Cash advance`",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=ReplyKeyboardRemove(),
        )

    # ── AUDIT ──
    elif data == "show_audit":
        await query.message.reply_text("⚠️ *Audit*\nChoose an option:", parse_mode=ParseMode.MARKDOWN, reply_markup=audit_menu())

    elif data == "audit_run":
        open_flags = [f for f in session["flags"] if not f.get("resolved")]
        records = session["records"]
        totals = (
            f"\n📊 *TOTALS:*\n"
            f"  💸 Deductions: ${sum(r['amount'] for r in records if r['payment_type']=='deduction'):,.2f}\n"
            f"  💚 Reimbursements: ${sum(r['amount'] for r in records if r['payment_type']=='reimbursement'):,.2f}\n"
            f"  ⭐ Bonuses: ${sum(r['amount'] for r in records if r['payment_type']=='bonus'):,.2f}\n"
            f"  🔴 Wage Holds: ${sum(r['amount'] for r in records if r['payment_type']=='wagehold'):,.2f}\n"
            f"  📋 Total Records: {len(records)}"
        )
        if not open_flags:
            msg = "✅ *PRE-PAYMENT AUDIT*\n\nNo active flags.\n*Cleared to process payments.*\n" + totals
        else:
            items = "\n".join(f"  • `[{f['id']}]` *{f['category']}* — {f['note']}" for f in open_flags)
            msg = (f"⚠️ *PRE-PAYMENT AUDIT*\n\n🔴 *{len(open_flags)} UNRESOLVED FLAG(S)*\n\n{items}\n\n"
                   f"❗ *Do NOT process until all flags resolved.*\n"
                   f"Use `/resolve [ID]` to clear.\n" + totals)
        await query.message.reply_text(msg, parse_mode=ParseMode.MARKDOWN, reply_markup=audit_menu())

    elif data == "audit_flags":
        summary = build_flags_summary(session)
        open_count = len([f for f in session["flags"] if not f.get("resolved")])
        resolved_count = len([f for f in session["flags"] if f.get("resolved")])
        await query.message.reply_text(
            summary + f"\n\n📊 *{open_count} open* | ✅ *{resolved_count} resolved*",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=audit_menu(),
        )

    # ── EXPORT MENU ──
    elif data in ("show_export", "export_menu"):
        await query.message.reply_text("📊 *Export Excel*\nFilter by:", parse_mode=ParseMode.MARKDOWN, reply_markup=export_menu())

    # ── EXPORT BY NAME ──
    elif data == "export_by_name":
        if not session["records"]:
            await query.message.reply_text("⚠️ No records yet.", reply_markup=export_menu())
        else:
            await query.message.reply_text("👤 *Export by Name*\nSelect a driver:", parse_mode=ParseMode.MARKDOWN, reply_markup=name_filter_menu(session))

    elif data.startswith("fname_"):
        value = data.replace("fname_", "")
        session["export_filter"] = {"by": "name", "value": value}
        label = "All Drivers" if value == "all" else value
        await query.message.reply_text(f"📅 Select period for *{label}*:", parse_mode=ParseMode.MARKDOWN, reply_markup=period_menu("pname"))

    elif data.startswith("pname_"):
        period = data.replace("pname_", "")
        ef = session["export_filter"]
        records = filter_records(session["records"], by=ef.get("by"), value=ef.get("value"), period=period if period != "all" else None)
        label = f"Name: {ef.get('value', 'All')} | {period.replace('_', ' ').title()}"
        await send_excel(query.message, records, label)

    # ── EXPORT BY UNIT ──
    elif data == "export_by_unit":
        if not session["records"]:
            await query.message.reply_text("⚠️ No records yet.", reply_markup=export_menu())
        else:
            await query.message.reply_text("🚛 *Export by Unit*\nSelect a unit:", parse_mode=ParseMode.MARKDOWN, reply_markup=unit_filter_menu(session))

    elif data.startswith("funit_"):
        value = data.replace("funit_", "")
        session["export_filter"] = {"by": "unit", "value": value}
        label = "All Units" if value == "all" else f"Unit {value}"
        await query.message.reply_text(f"📅 Select period for *{label}*:", parse_mode=ParseMode.MARKDOWN, reply_markup=period_menu("punit"))

    elif data.startswith("punit_"):
        period = data.replace("punit_", "")
        ef = session["export_filter"]
        records = filter_records(session["records"], by=ef.get("by"), value=ef.get("value"), period=period if period != "all" else None)
        label = f"Unit: {ef.get('value', 'All')} | {period.replace('_', ' ').title()}"
        await send_excel(query.message, records, label)

    # ── EXPORT BY TYPE ──
    elif data == "export_by_type":
        await query.message.reply_text("📋 *Export by Type*\nSelect a type:", parse_mode=ParseMode.MARKDOWN, reply_markup=type_filter_menu())

    elif data.startswith("ftype_"):
        value = data.replace("ftype_", "")
        session["export_filter"] = {"by": "type", "value": value}
        label = "All Types" if value == "all" else value.title()
        await query.message.reply_text(f"📅 Select period for *{label}*:", parse_mode=ParseMode.MARKDOWN, reply_markup=period_menu("ptype"))

    elif data.startswith("ptype_"):
        period = data.replace("ptype_", "")
        ef = session["export_filter"]
        records = filter_records(session["records"], by=ef.get("by"), value=ef.get("value"), period=period if period != "all" else None)
        label = f"Type: {ef.get('value', 'All').title()} | {period.replace('_', ' ').title()}"
        await send_excel(query.message, records, label)

    # ── EXPORT BY PERIOD ONLY ──
    elif data == "export_by_period":
        session["export_filter"] = {"by": None, "value": None}
        await query.message.reply_text("📅 *Export by Period*\nSelect a period:", parse_mode=ParseMode.MARKDOWN, reply_markup=period_menu("ponly"))

    elif data.startswith("ponly_"):
        period = data.replace("ponly_", "")
        records = filter_records(session["records"], period=period if period != "all" else None)
        label = period.replace("_", " ").title()
        await send_excel(query.message, records, label)


# ── Main message handler ──────────────────────────────────────────────────────

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    session = get_session(update.effective_user.id)
    text = update.message.text.strip()

    # ── Persistent keyboard button taps ──
    if text == "📋 Records":
        await update.message.reply_text("📋 *Records*\nChoose a category:", parse_mode=ParseMode.MARKDOWN, reply_markup=records_menu())
        return
    elif text == "➕ Add New":
        await update.message.reply_text("➕ *Add New Record*\nChoose type:", parse_mode=ParseMode.MARKDOWN, reply_markup=add_new_menu())
        return
    elif text == "⚠️ Audit":
        await update.message.reply_text("⚠️ *Audit*\nChoose an option:", parse_mode=ParseMode.MARKDOWN, reply_markup=audit_menu())
        return
    elif text == "📊 Export":
        await update.message.reply_text("📊 *Export Excel*\nFilter by:", parse_mode=ParseMode.MARKDOWN, reply_markup=export_menu())
        return

    # ── Awaiting new record input ──
    if session.get("awaiting") and session["awaiting"].startswith("new_record_"):
        ptype = session["awaiting"].replace("new_record_", "")
        record = parse_new_record(text)
        if not record:
            await update.message.reply_text(
                "❌ Wrong format. Use:\n`Unit - Name - Amount - Type - PaymentMethod - Note`\n\nExample:\n`2154 - Jama Ahmed - 2500 - deduction - Check - Cash advance`",
                parse_mode=ParseMode.MARKDOWN,
            )
            return

        record["payment_type"] = ptype
        rid = next_record_id(session)
        now = datetime.now()
        record.update({"id": rid, "datetime": now, "date": now.strftime("%m/%d/%Y %H:%M")})
        session["records"].append(record)
        session["awaiting"] = None

        if record["payment_type"].lower() == "wagehold":
            add_flag(session, "CARRIER HOLD", f"Wage hold — {record['name']} ${record['amount']:,.2f}")

        emoji_map = {"deduction": "💸", "reimbursement": "💚", "bonus": "⭐", "wagehold": "🔴"}
        emoji = emoji_map.get(record["payment_type"].lower(), "📄")

        await update.message.reply_text(
            f"{emoji} *{record['payment_type'].title()} Added* `{rid}`\n\n"
            f"🚛 *Unit:* {record['unit']}\n"
            f"👤 *Driver:* {record['name']}\n"
            f"💰 *Amount:* ${record['amount']:,.2f}\n"
            f"📋 *Type:* {record['payment_type'].title()}\n"
            f"💳 *Method:* {record['payment_method']}\n"
            f"📝 *Note:* {record['note']}\n"
            f"📅 *Week:* {get_week_label(now)}",
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=MAIN_KEYBOARD,
        )
        return

    # ── Gemini AI fallback ──
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action="typing")
    reply = call_gemini(session, text)

    for kw, category in FLAG_KEYWORDS.items():
        if kw in reply:
            note = text[:120] + ("..." if len(text) > 120 else "")
            fid = add_flag(session, category, note)
            reply = reply.replace(kw, f"{kw} `[{fid}]`", 1)

    try:
        await update.message.reply_text(reply[:4000], parse_mode=ParseMode.MARKDOWN, reply_markup=MAIN_KEYBOARD)
    except Exception:
        await update.message.reply_text(reply[:4000], reply_markup=MAIN_KEYBOARD)


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise ValueError("TELEGRAM_BOT_TOKEN not set.")
    if not os.environ.get("GEMINI_API_KEY"):
        raise ValueError("GEMINI_API_KEY not set.")

    app = Application.builder().token(token).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("resolve", cmd_resolve))
    app.add_handler(CallbackQueryHandler(button_callback))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    logger.info("LogiCheck bot starting (Gemini + Buttons)...")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
